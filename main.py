import os
import pdfplumber
import requests
import json
from pathlib import Path
from InquirerPy import inquirer
from tkinter import Tk, filedialog
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

from data_models import Student, HistoryEntry
from constants import CMD_NULL, CMD_FILTER, CMD_FINISH, FILTER_DIR_PATH, LOCAL_VERSION_FILE


class GradingApp:
    """CLI tool for entering grades into an Excel student list."""

    def __init__(self):
        self.students: dict[str, Student] = {}
        self.history: list[HistoryEntry] = []

        self.filters: dict[str, list[str]] = {}
        self.active_filter: str | None = None
        self.file_path: str

        self.min_grade: float
        self.max_grade: float

        self.number_col_letter: str
        self.number_row: int
        self.number_col: int
        self.name_col_letter: str
        self.name_row: int
        self.name_col: int
        self.grade_col_letter: str
        self.grade_row: int
        self.grade_col: int


    def main(self) -> None:
        """Run the full grading workflow."""
        self.clear_screen()

        self.prompt_cell_references()
        self.fetch_student_records()
        self.setup_input_structure()
        self.load_filters()

        self.grading()
        self.save_output_file()


    def prompt_cell_references(self) -> None:
        self.file_path = self.open_input_file()

        student_number_idx = input("Enter the cell reference where student numbers start (e.g., A2): ").strip().upper()
        student_name_idx = input("Enter the cell reference where student names start (e.g., B2): ").strip().upper()
        student_grade_idx = input("Enter the cell reference where student grades will start (e.g., C2): ").strip().upper()

        self.number_col_letter, self.number_row = coordinate_from_string(student_number_idx)
        self.name_col_letter, self.name_row = coordinate_from_string(student_name_idx)
        self.grade_col_letter, self.grade_row = coordinate_from_string(student_grade_idx)
        if self.number_row != self.name_row:
            raise ValueError("Student number and student name must start on the same row.")
        if self.number_row != self.grade_row:
            raise ValueError("Student number and student grade must start on the same row.")
        self.number_col = column_index_from_string(self.number_col_letter)
        self.name_col = column_index_from_string(self.name_col_letter)
        self.grade_col = column_index_from_string(self.grade_col_letter)


    def fetch_student_records(self) -> None:
        """Load students and existing grades from Excel."""

        wb = load_workbook(self.file_path)
        ws = wb.active

        if not ws:
            raise ValueError("Excel file does not contain any sheets.")

        for row_idx in range(self.number_row, ws.max_row + 1):
            student_id = ws.cell(row=row_idx, column=self.number_col).value
            student_name = ws.cell(row=row_idx, column=self.name_col).value
            student_grade = ws.cell(row=row_idx, column=self.grade_col).value

            if student_id is None and student_name is None:
                # Stop reading at the first completely empty row.
                break

            try:
                if isinstance(student_grade, str):
                    student_grade = student_grade.strip()

                    if student_grade == "":
                        student_grade = None
                    else:
                        # Support both comma and dot as decimal separators.
                        student_grade = student_grade.replace(",", ".")
                        student_grade = float(student_grade)

                elif student_grade is not None:
                    student_grade = float(student_grade)

            except (TypeError, ValueError):
                raise ValueError(
                    f"Invalid existing grade '{student_grade}' "
                    f"for student '{student_id}' at row {row_idx}."
                )

            self.students[str(student_id)] = Student(
                str(student_id),
                str(student_name),
                student_grade
            )
    

    def load_filters(self) -> None:
        """Load optional student ID filters from PDF and TXT files."""

        if not os.path.exists(FILTER_DIR_PATH):
            self.clear_screen()
            input(
                f"Filter directory '{FILTER_DIR_PATH}' not found. "
                "Continuing without filters (highly recommended for faster grading)..."
            )
            return
        
        filters = os.listdir(FILTER_DIR_PATH)
        if not filters:
            self.clear_screen()
            input(
                f"No filters found. "
                "Continuing without filters (highly recommended for faster grading)..."
            )
            return

        for filename in filters:
            file_path = os.path.join(FILTER_DIR_PATH, filename)
            filter_name, extension = os.path.splitext(filename)
            extension = extension.lower()

            if extension not in [".pdf", ".txt"]:
                continue

            try:
                if extension == ".pdf":
                    text_all = []

                    with pdfplumber.open(file_path) as pdf:
                        for page in pdf.pages:
                            text = page.extract_text()
                            if text:
                                text_all.append(text)

                    text = "\n".join(text_all)

                elif extension == ".txt":
                    with open(file_path, "r", encoding="utf-8") as file:
                        text = file.read()

                tokens = text.split()

                # Extract numeric tokens as student IDs.
                # TXT files may contain names or other text, but only pure numbers are used.
                numbers = [token for token in tokens if token.isdigit()]

                self.filters[filter_name] = numbers

            except Exception as exc:
                print(f"Error reading filter file '{file_path}': {exc}")
                self.filters[filter_name] = []


    def setup_input_structure(self) -> None:
        """Ask user for allowed grade range."""

        while True:
            try:
                # Support both comma and dot as decimal separators.
                min_str = input("Enter the minimum possible grade: ").strip().replace(",", ".")
                max_str = input("Enter the maximum possible grade: ").strip().replace(",", ".")

                self.min_grade = float(min_str)
                self.max_grade = float(max_str)

                if self.min_grade >= self.max_grade:
                    print("Minimum grade must be less than maximum grade. Please try again.")
                    continue

                break

            except ValueError:
                print("Invalid input! Please enter numeric values (e.g., 0, 100, 0.0, 10.5).")


    def grading(self) -> None:
        """Interactive grading loop."""

        while True:
            self.clear_screen()
            self.print_history()
            remaining_students = [
                {"name": f"{student_id}   {student.student_name}", "value": student_id}
                for student_id, student in self.students.items()
                if student.grade is None
            ]
            if self.active_filter is not None:
                filtered_students = [
                    student
                    for student in remaining_students
                    if student["value"] in self.filters.get(self.active_filter, [])
                ]
            else:
                filtered_students = remaining_students

            print(f"Remaining: {len(remaining_students)}/{len(self.students)}")
            selection = inquirer.fuzzy(
                message = "Select a student:",
                choices = (
                    [{'name': CMD_NULL, 'value': CMD_NULL}]
                    + filtered_students 
                    + 
                    [{'name': CMD_FILTER, 'value': CMD_FILTER},
                     {'name': CMD_FINISH, 'value': CMD_FINISH}]
                )
            ).execute()
            
            if selection == CMD_NULL:
                # null choice to prevent accidental selections
                continue
            elif selection == CMD_FILTER:
                self.change_filter()
            elif selection == CMD_FINISH:
                break
            else:
                grade = self.get_grade()
                if grade is not None:
                    self.assign_grade(selection, grade)


    def get_grade(self) -> float | None:
        """Prompt user for a valid grade or allow cancel."""

        while True:
            grade_str = input("Enter grade (q to cancel): ").strip()

            if grade_str.lower() in ["q", "exit", "cancel", "break"]:
                return
            
            # Support both comma and dot as decimal separators.
            grade_str = grade_str.replace(",", ".")
            try:
                grade = float(grade_str)
            except ValueError:
                print("Invalid input! Please enter a numeric value.")
                continue

            if not (self.min_grade <= grade <= self.max_grade):
                print(f"Grade must be between {self.min_grade} and {self.max_grade}.")
                continue

            return grade


    def assign_grade(self, student_id: str, grade: float) -> None:
        self.students[student_id].grade = grade
        self.history.append(HistoryEntry(student_id, self.students[student_id].student_name, grade))


    def save_output_file(self) -> None:
        """Write grades back to the Excel file."""

        wb = load_workbook(self.file_path)
        ws = wb.active

        if not ws:
            raise ValueError("Excel file does not contain any sheets.")

        for row_idx in range(self.grade_row, ws.max_row + 1):
            student_id = ws.cell(row=row_idx, column=self.number_col).value
            if student_id is None:
                continue
            student_id = str(student_id)
            
            if student_id not in self.students:
                continue
            
            student = self.students[student_id]
            if student.grade is not None:
                ws.cell(row=row_idx, column=self.grade_col).value = student.grade

        while True:
            try:
                wb.save(self.file_path)
                return
            except PermissionError:
                self.clear_screen()
                input(
                    "\nCannot save Excel file.\n"
                    "Close the file in Excel and press ENTER to retry..."
                )


    def change_filter(self) -> None:
        selection = inquirer.select(
            message="Select a filter to change:",
            choices= ["None"] + list(self.filters.keys())
        ).execute()

        if selection == "None":
            self.active_filter = None
        else:
            self.active_filter = selection


    def print_history(self) -> None:
        for entry in self.history:
            print(f"{format(f'{entry.student_id}   {entry.student_name}', '<35')}{entry.grade}")
        print()


    def clear_screen(self) -> None:
        os.system("cls" if os.name == "nt" else "clear")


    def open_input_file(self) -> str:
        """Open file picker for Excel input file."""

        input("Select student excel file...")
        root = Tk()
        root.withdraw()  

        file_path = filedialog.askopenfilename(
            title="Select student Excel file",
            filetypes=[
                ("Excel files", "*.xlsx"),
            ],
        )
        root.destroy()

        if not file_path:
            raise FileNotFoundError("No input file selected!")

        return file_path


def check_updates():
    print("Starting...")

    local_file = Path(__file__).parent / LOCAL_VERSION_FILE
    if not local_file.exists():
        return
    local_version = tuple(map(int, json.loads(local_file.read_text())["version"].split(".")))
    remote_version_data = requests.get(
        "https://raw.githubusercontent.com/oguz-duman/grading/main/latest.json",
        timeout=3
    ).json()
    remote_version = tuple(map(int, remote_version_data["version"].split(".")))

    if remote_version > local_version:
        os.system("cls" if os.name == "nt" else "clear")
        print(f"\nUpdate available ({'.'.join(map(str, local_version))} → {remote_version_data['version']})")
        if "notes" in remote_version_data:
            print(remote_version_data["notes"])
        print("You may run: git pull\n")
        input("...")


if __name__ == "__main__":
    try:
        check_updates()
    except Exception:
        pass

    try:
        grading_app = GradingApp()
        grading_app.main()
    except FileNotFoundError as exc:
        print(f"FileNotFoundError: {exc}")
    except Exception as exc:
        print(f"Unexpected error: {exc}")
